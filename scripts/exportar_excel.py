"""
exportar_excel.py
Exporta todas las tablas de la DB activos_ml a un archivo Excel (.xlsx).

Una hoja por tabla + hoja "Descripcion" con:
    - Origen de cada variable (FUENTE = dato externo / CALCULADO = derivado)
    - Descripcion de que representa
    - Formula o metodo de calculo (si aplica)
    - Condiciones o umbrales usados

Uso:
    python scripts/exportar_excel.py
"""

import sys
import os
from datetime import datetime

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.data.database import query_df

# ─────────────────────────────────────────────────────────────
# Tablas a exportar (en orden de aparicion en el proyecto)
# ─────────────────────────────────────────────────────────────

TABLAS = [
    ("activos",                   "Universo de activos"),
    ("precios_diarios",           "Precios OHLCV diarios"),
    ("indicadores_tecnicos",      "Indicadores tecnicos calculados"),
    ("scoring_tecnico",           "Scoring rule-based"),
    ("operaciones_backtest",      "Operaciones simuladas backtest"),
    ("resultados_backtest",       "Resumen de resultados backtest"),
    ("features_sector",           "Features sectoriales (Z-Scores)"),
    ("features_ml",               "Feature store ML (train/test/bt)"),
    ("resultados_modelos_ml",     "Metricas de modelos ML"),
    ("modelos_produccion",        "Plan de despliegue por modelo"),
    ("features_precio_accion",    "Features estructura precio/volumen"),
    ("features_market_structure", "Features estructura de mercado"),
    ("resultados_ml_filter",      "Backtest con filtro ML"),
    ("log_ejecuciones",           "Log de ejecuciones de scripts"),
]


# ─────────────────────────────────────────────────────────────
# Diccionario de datos — Descripcion de variables
# ─────────────────────────────────────────────────────────────
# Columnas: tabla | columna | tipo_bd | origen | descripcion | formula | condiciones

DESC = [

    # ── activos ───────────────────────────────────────────────────────
    ("activos","id","SERIAL","CALCULADO","Identificador interno autoincremental","SERIAL (secuencia PostgreSQL)","Unico, primario"),
    ("activos","ticker","VARCHAR","FUENTE","Simbolo bursatil del activo","Dato externo (configuracion manual)","19 tickers definidos en config.py"),
    ("activos","nombre","VARCHAR","FUENTE","Nombre completo de la empresa","Dato externo (configuracion manual)",""),
    ("activos","sector","VARCHAR","FUENTE","Sector economico de clasificacion","Dato externo (GICS aproximado)","Financials / Consumer Staples / Consumer Discretionary / Automotive"),
    ("activos","activo","BOOLEAN","CALCULADO","Flag: activo en el universo de seguimiento","Valor por defecto TRUE",""),
    ("activos","created_at","TIMESTAMP","CALCULADO","Fecha/hora de insercion del registro","NOW() al insertar",""),

    # ── precios_diarios ───────────────────────────────────────────────
    ("precios_diarios","id","SERIAL","CALCULADO","Identificador interno autoincremental","SERIAL",""),
    ("precios_diarios","ticker","VARCHAR","FUENTE","Simbolo bursatil","API externa (FMP / Alpha Vantage)",""),
    ("precios_diarios","fecha","DATE","FUENTE","Fecha de la sesion bursatil","API externa","Solo dias habiles de mercado"),
    ("precios_diarios","open","NUMERIC","FUENTE","Precio de apertura de la sesion","API externa","USD, ajustado por splits"),
    ("precios_diarios","high","NUMERIC","FUENTE","Precio maximo intradía","API externa","USD"),
    ("precios_diarios","low","NUMERIC","FUENTE","Precio minimo intradía","API externa","USD"),
    ("precios_diarios","close","NUMERIC","FUENTE","Precio de cierre de la sesion","API externa","USD"),
    ("precios_diarios","volume","BIGINT","FUENTE","Volumen de acciones transaccionadas","API externa","Unidades enteras"),
    ("precios_diarios","adj_close","NUMERIC","FUENTE","Precio de cierre ajustado por dividendos y splits","API externa","Puede diferir de close historico"),
    ("precios_diarios","created_at","TIMESTAMP","CALCULADO","Fecha/hora de insercion","NOW()",""),

    # ── indicadores_tecnicos ──────────────────────────────────────────
    ("indicadores_tecnicos","ticker","VARCHAR","FUENTE","Simbolo bursatil","Igual que precios_diarios",""),
    ("indicadores_tecnicos","fecha","DATE","FUENTE","Fecha de calculo","Igual que precios_diarios",""),
    ("indicadores_tecnicos","sma21","NUMERIC","CALCULADO","Media movil simple de 21 dias","SMA(close, 21) = sum(close[i-20..i]) / 21","Requiere min 21 barras (warmup)"),
    ("indicadores_tecnicos","sma50","NUMERIC","CALCULADO","Media movil simple de 50 dias","SMA(close, 50)","Requiere min 50 barras"),
    ("indicadores_tecnicos","sma200","NUMERIC","CALCULADO","Media movil simple de 200 dias","SMA(close, 200)","Requiere min 200 barras"),
    ("indicadores_tecnicos","dist_sma21","NUMERIC","CALCULADO","Distancia porcentual del precio a SMA21","(close - sma21) / sma21 * 100","Positivo = precio sobre la media"),
    ("indicadores_tecnicos","dist_sma50","NUMERIC","CALCULADO","Distancia porcentual del precio a SMA50","(close - sma50) / sma50 * 100",""),
    ("indicadores_tecnicos","dist_sma200","NUMERIC","CALCULADO","Distancia porcentual del precio a SMA200","(close - sma200) / sma200 * 100",""),
    ("indicadores_tecnicos","rsi14","NUMERIC","CALCULADO","Indice de Fuerza Relativa de 14 periodos","RSI = 100 - 100/(1 + RS); RS = avg_gain_14 / avg_loss_14","Rango 0-100. Wilder smoothing"),
    ("indicadores_tecnicos","momentum","NUMERIC","CALCULADO","Momentum de precio en 14 dias","close - close.shift(14)","Positivo = tendencia alcista"),
    ("indicadores_tecnicos","macd","NUMERIC","CALCULADO","Moving Average Convergence Divergence","EMA(close,12) - EMA(close,26)","EMA = Exponential Moving Average"),
    ("indicadores_tecnicos","macd_signal","NUMERIC","CALCULADO","Linea de senal del MACD","EMA(macd, 9)","Suavizado del MACD"),
    ("indicadores_tecnicos","macd_hist","NUMERIC","CALCULADO","Histograma MACD (diferencia)","macd - macd_signal","Positivo = momento alcista"),
    ("indicadores_tecnicos","atr14","NUMERIC","CALCULADO","Average True Range de 14 dias","ATR = Wilder_avg(True_Range, 14); TR = max(high-low, |high-prev_close|, |low-prev_close|)","Mide volatilidad absoluta en USD"),
    ("indicadores_tecnicos","bb_upper","NUMERIC","CALCULADO","Banda de Bollinger superior","SMA(close,20) + 2*std(close,20)","2 desviaciones estandar sobre la media"),
    ("indicadores_tecnicos","bb_middle","NUMERIC","CALCULADO","Banda de Bollinger media","SMA(close, 20)","Media movil simple 20 dias"),
    ("indicadores_tecnicos","bb_lower","NUMERIC","CALCULADO","Banda de Bollinger inferior","SMA(close,20) - 2*std(close,20)","2 desviaciones estandar bajo la media"),
    ("indicadores_tecnicos","obv","NUMERIC","CALCULADO","On Balance Volume acumulado","OBV[i] = OBV[i-1] + volume si close>prev_close; - volume si close<prev_close","Acumulado desde inicio de la serie"),
    ("indicadores_tecnicos","vol_relativo","NUMERIC","CALCULADO","Volumen relativo vs promedio 20 dias","volume / SMA(volume, 20)","1.0 = volumen normal; >2.0 = pico"),
    ("indicadores_tecnicos","adx","NUMERIC","CALCULADO","Average Directional Index de 14 dias","ADX = Wilder_avg(|DI+ - DI-| / (DI+ + DI-), 14)","Rango 0-100; >25 = tendencia definida"),

    # ── scoring_tecnico ───────────────────────────────────────────────
    ("scoring_tecnico","ticker","VARCHAR","FUENTE","Simbolo bursatil","Igual que precios_diarios",""),
    ("scoring_tecnico","fecha","DATE","FUENTE","Fecha de calculo","Igual que precios_diarios",""),
    ("scoring_tecnico","cond_rsi","BOOLEAN","CALCULADO","Condicion RSI no extremo","RSI14 > 35 AND RSI14 < 65","True = ni sobrevendido ni sobrecomprado"),
    ("scoring_tecnico","cond_macd","BOOLEAN","CALCULADO","Condicion MACD positivo","macd_hist > 0","True = momentum alcista"),
    ("scoring_tecnico","cond_sma21","BOOLEAN","CALCULADO","Precio sobre SMA21","close > sma21","True = tendencia corto plazo alcista"),
    ("scoring_tecnico","cond_sma50","BOOLEAN","CALCULADO","Precio sobre SMA50","close > sma50","True = tendencia mediano plazo alcista"),
    ("scoring_tecnico","cond_sma200","BOOLEAN","CALCULADO","Precio sobre SMA200","close > sma200","True = tendencia largo plazo alcista"),
    ("scoring_tecnico","cond_momentum","BOOLEAN","CALCULADO","Momentum positivo","momentum > 0","True = precio mas alto que hace 14 dias"),
    ("scoring_tecnico","score_ponderado","NUMERIC","CALCULADO","Score compuesto ponderado entre 0 y 1","RSI*0.20 + MACD*0.20 + SMA21*0.10 + SMA50*0.15 + SMA200*0.20 + MOM*0.15","Suma de condiciones True * peso de cada condicion"),
    ("scoring_tecnico","condiciones_ok","SMALLINT","CALCULADO","Cantidad de condiciones True (0-6)","sum(cond_rsi, cond_macd, cond_sma21, cond_sma50, cond_sma200, cond_momentum)","Entero entre 0 y 6"),
    ("scoring_tecnico","senal","VARCHAR","CALCULADO","Senal generada por el scoring","LONG si score_ponderado >= 0.60; NEUTRAL en caso contrario","Umbral SCORE_ENTRADA_UMBRAL = 0.60"),

    # ── operaciones_backtest ──────────────────────────────────────────
    ("operaciones_backtest","estrategia_entrada","VARCHAR","CALCULADO","Estrategia de entrada usada (E1-E4)","E1=score>=0.60 | E2=score>=0.60 + vol_spike | E3=score>=0.60 + macd_hist_creciente | E4=score>=0.50","Combinado con estrategia_salida en 4x4"),
    ("operaciones_backtest","estrategia_salida","VARCHAR","CALCULADO","Estrategia de salida usada (S1-S4)","S1=timeout_20d | S2=score<0.30 | S3=macd_hist<0 | S4=trailing_stop_3pct",""),
    ("operaciones_backtest","ticker","VARCHAR","FUENTE","Simbolo bursatil de la operacion","Igual que precios_diarios",""),
    ("operaciones_backtest","segmento","VARCHAR","CALCULADO","Segmento temporal de la operacion","TRAIN (70%) / TEST (15%) / BACKTEST (15%)","Split cronologico por ticker"),
    ("operaciones_backtest","fecha_entrada","DATE","CALCULADO","Fecha en que se simulo la entrada","Dia en que se activo la senal LONG",""),
    ("operaciones_backtest","precio_entrada","NUMERIC","FUENTE","Precio de cierre al entrar","close en fecha_entrada",""),
    ("operaciones_backtest","score_entrada","NUMERIC","CALCULADO","Score ponderado al momento de entrada","scoring_tecnico.score_ponderado en fecha_entrada",""),
    ("operaciones_backtest","fecha_salida","DATE","CALCULADO","Fecha en que se simulo la salida","Segun condicion de salida o timeout",""),
    ("operaciones_backtest","precio_salida","NUMERIC","FUENTE","Precio de cierre al salir","close en fecha_salida",""),
    ("operaciones_backtest","motivo_salida","VARCHAR","CALCULADO","Por que se cerro la posicion","TIMEOUT / STOP_LOSS / TAKE_PROFIT / SENAL",""),
    ("operaciones_backtest","dias_posicion","INTEGER","CALCULADO","Cantidad de dias mantenida la posicion","fecha_salida - fecha_entrada","Maximo 20 dias (TIMEOUT_DIAS)"),
    ("operaciones_backtest","retorno_pct","NUMERIC","CALCULADO","Retorno de la operacion en porcentaje","(precio_salida - precio_entrada) / precio_entrada * 100","Bruto, sin costos de transaccion"),
    ("operaciones_backtest","resultado","VARCHAR","CALCULADO","Clasificacion del resultado","GANANCIA si retorno > 1% | PERDIDA si retorno < -1% | NEUTRO en otro caso","Umbral UMBRAL_NEUTRO = 0.01"),

    # ── resultados_backtest ───────────────────────────────────────────
    ("resultados_backtest","estrategia_entrada","VARCHAR","CALCULADO","Estrategia de entrada (E1-E4)","Agrupacion de operaciones_backtest",""),
    ("resultados_backtest","estrategia_salida","VARCHAR","CALCULADO","Estrategia de salida (S1-S4)","Agrupacion de operaciones_backtest",""),
    ("resultados_backtest","ticker","VARCHAR","FUENTE","Ticker (NULL = resultado agregado del scope)","","NULL = todos los tickers del segmento"),
    ("resultados_backtest","segmento","VARCHAR","CALCULADO","TRAIN / TEST / BACKTEST","",""),
    ("resultados_backtest","total_operaciones","INTEGER","CALCULADO","Cantidad de operaciones simuladas","COUNT(*) de operaciones_backtest",""),
    ("resultados_backtest","ganancias","INTEGER","CALCULADO","Operaciones con resultado GANANCIA","COUNT WHERE resultado='GANANCIA'","retorno > +1%"),
    ("resultados_backtest","perdidas","INTEGER","CALCULADO","Operaciones con resultado PERDIDA","COUNT WHERE resultado='PERDIDA'","retorno < -1%"),
    ("resultados_backtest","neutros","INTEGER","CALCULADO","Operaciones con resultado NEUTRO","COUNT WHERE resultado='NEUTRO'","retorno entre -1% y +1%"),
    ("resultados_backtest","win_rate","NUMERIC","CALCULADO","Tasa de acierto","ganancias / total_operaciones","Entre 0 y 1"),
    ("resultados_backtest","retorno_promedio_pct","NUMERIC","CALCULADO","Retorno promedio por operacion","mean(retorno_pct)","En porcentaje"),
    ("resultados_backtest","retorno_total_pct","NUMERIC","CALCULADO","Retorno acumulado total","sum(retorno_pct)","Suma simple, no compuesta"),
    ("resultados_backtest","max_drawdown_pct","NUMERIC","CALCULADO","Maximo drawdown del equity","max(peak - trough) sobre la curva de retornos acumulados","Mide el peor momento de la estrategia"),
    ("resultados_backtest","profit_factor","NUMERIC","CALCULADO","Factor de beneficio","sum(ganancias) / |sum(perdidas)|","Mayor a 1 = estrategia rentable"),
    ("resultados_backtest","dias_promedio_posicion","NUMERIC","CALCULADO","Duracion media de las posiciones","mean(dias_posicion)","Dias habiles de mercado"),

    # ── features_sector ───────────────────────────────────────────────
    ("features_sector","ticker","VARCHAR","FUENTE","Simbolo bursatil","",""),
    ("features_sector","fecha","DATE","FUENTE","Fecha","",""),
    ("features_sector","sector","VARCHAR","FUENTE","Sector de clasificacion del ticker","Igual que activos.sector",""),
    ("features_sector","z_rsi_sector","NUMERIC","CALCULADO","Z-Score de RSI14 relativo al sector en esa fecha","(rsi14 - mean_sector(rsi14)) / std_sector(rsi14)","Normalizado respecto a todos los tickers del mismo sector en la misma fecha"),
    ("features_sector","z_retorno_1d_sector","NUMERIC","CALCULADO","Z-Score del retorno diario relativo al sector","(retorno_1d - mean_sector) / std_sector","Retorno 1d = (close - prev_close) / prev_close * 100"),
    ("features_sector","z_retorno_5d_sector","NUMERIC","CALCULADO","Z-Score del retorno 5d relativo al sector","(retorno_5d - mean_sector) / std_sector","Retorno 5d = (close - close_5d_ago) / close_5d_ago * 100"),
    ("features_sector","z_vol_sector","NUMERIC","CALCULADO","Z-Score del volumen relativo al sector","(vol_relativo - mean_sector) / std_sector","vol_relativo = volumen / SMA(20) del volumen"),
    ("features_sector","z_dist_sma50_sector","NUMERIC","CALCULADO","Z-Score de dist_sma50 relativo al sector","(dist_sma50 - mean_sector) / std_sector",""),
    ("features_sector","z_adx_sector","NUMERIC","CALCULADO","Z-Score del ADX relativo al sector","(adx - mean_sector) / std_sector",""),
    ("features_sector","pct_long_sector","NUMERIC","CALCULADO","Porcentaje de tickers con senal LONG en el sector","count(senal='LONG') / count_total_sector","Mide amplitud del mercado (breadth)"),
    ("features_sector","rank_retorno_sector","INTEGER","CALCULADO","Ranking del ticker por retorno diario dentro del sector","rank() OVER (sector, fecha ORDER BY retorno_1d DESC)","1 = mejor rendimiento del dia"),
    ("features_sector","rsi_sector_avg","NUMERIC","CALCULADO","RSI promedio del sector en la fecha","mean(rsi14) de todos los tickers del sector","Contexto de momentum sectorial"),
    ("features_sector","adx_sector_avg","NUMERIC","CALCULADO","ADX promedio del sector en la fecha","mean(adx) de todos los tickers del sector","Contexto de tendencia sectorial"),
    ("features_sector","retorno_1d_sector_avg","NUMERIC","CALCULADO","Retorno diario promedio del sector","mean(retorno_1d) de todos los tickers del sector","Mide la direccion del sector ese dia"),

    # ── features_ml ───────────────────────────────────────────────────
    ("features_ml","ticker","VARCHAR","FUENTE","Simbolo bursatil","",""),
    ("features_ml","nombre","VARCHAR","FUENTE","Nombre de la empresa","Igual que activos.nombre",""),
    ("features_ml","sector","VARCHAR","FUENTE","Sector de clasificacion","Igual que activos.sector",""),
    ("features_ml","fecha","DATE","FUENTE","Fecha de la barra","",""),
    ("features_ml","segmento","VARCHAR","CALCULADO","Particion temporal: TRAIN / TEST / BACKTEST","70% primeras fechas -> TRAIN | 15% -> TEST | 15% ultimas -> BACKTEST","Split cronologico por ticker, NO aleatorio. Nunca hay contaminacion temporal"),
    ("features_ml","close","NUMERIC","FUENTE","Precio de cierre","Igual que precios_diarios.close",""),
    ("features_ml","vol_relativo","NUMERIC","CALCULADO","Volumen relativo vs SMA20","volume / SMA(volume,20)","Igual que indicadores_tecnicos.vol_relativo"),
    ("features_ml","rsi14","NUMERIC","CALCULADO","RSI 14 periodos","Ver indicadores_tecnicos.rsi14",""),
    ("features_ml","macd_hist","NUMERIC","CALCULADO","Histograma MACD","Ver indicadores_tecnicos.macd_hist",""),
    ("features_ml","dist_sma21","NUMERIC","CALCULADO","Distancia % a SMA21","Ver indicadores_tecnicos",""),
    ("features_ml","dist_sma50","NUMERIC","CALCULADO","Distancia % a SMA50","Ver indicadores_tecnicos",""),
    ("features_ml","dist_sma200","NUMERIC","CALCULADO","Distancia % a SMA200","Ver indicadores_tecnicos",""),
    ("features_ml","adx","NUMERIC","CALCULADO","ADX 14 periodos","Ver indicadores_tecnicos",""),
    ("features_ml","atr14","NUMERIC","CALCULADO","ATR 14 periodos (volatilidad)","Ver indicadores_tecnicos",""),
    ("features_ml","momentum","NUMERIC","CALCULADO","Momentum precio 14d","close - close.shift(14)",""),
    ("features_ml","bb_upper","NUMERIC","CALCULADO","Banda Bollinger superior","Ver indicadores_tecnicos",""),
    ("features_ml","bb_middle","NUMERIC","CALCULADO","Banda Bollinger media","Ver indicadores_tecnicos",""),
    ("features_ml","bb_lower","NUMERIC","CALCULADO","Banda Bollinger inferior","Ver indicadores_tecnicos",""),
    ("features_ml","score_ponderado","NUMERIC","CALCULADO","Score rule-based ponderado","Ver scoring_tecnico",""),
    ("features_ml","condiciones_ok","SMALLINT","CALCULADO","Condiciones rule-based activas","Ver scoring_tecnico",""),
    ("features_ml","cond_rsi","BOOLEAN","CALCULADO","Condicion RSI activa","Ver scoring_tecnico",""),
    ("features_ml","cond_macd","BOOLEAN","CALCULADO","Condicion MACD activa","Ver scoring_tecnico",""),
    ("features_ml","cond_sma21","BOOLEAN","CALCULADO","Condicion SMA21 activa","Ver scoring_tecnico",""),
    ("features_ml","cond_sma50","BOOLEAN","CALCULADO","Condicion SMA50 activa","Ver scoring_tecnico",""),
    ("features_ml","cond_sma200","BOOLEAN","CALCULADO","Condicion SMA200 activa","Ver scoring_tecnico",""),
    ("features_ml","cond_momentum","BOOLEAN","CALCULADO","Condicion momentum activa","Ver scoring_tecnico",""),
    ("features_ml","z_rsi_sector","NUMERIC","CALCULADO","Z-Score RSI vs sector","Ver features_sector",""),
    ("features_ml","z_retorno_1d_sector","NUMERIC","CALCULADO","Z-Score retorno 1d vs sector","Ver features_sector",""),
    ("features_ml","z_retorno_5d_sector","NUMERIC","CALCULADO","Z-Score retorno 5d vs sector","Ver features_sector",""),
    ("features_ml","z_vol_sector","NUMERIC","CALCULADO","Z-Score volumen vs sector","Ver features_sector",""),
    ("features_ml","z_dist_sma50_sector","NUMERIC","CALCULADO","Z-Score dist SMA50 vs sector","Ver features_sector",""),
    ("features_ml","z_adx_sector","NUMERIC","CALCULADO","Z-Score ADX vs sector","Ver features_sector",""),
    ("features_ml","pct_long_sector","NUMERIC","CALCULADO","% tickers LONG en el sector","Ver features_sector",""),
    ("features_ml","rank_retorno_sector","INTEGER","CALCULADO","Ranking retorno diario en sector","Ver features_sector",""),
    ("features_ml","rsi_sector_avg","NUMERIC","CALCULADO","RSI promedio del sector","Ver features_sector",""),
    ("features_ml","adx_sector_avg","NUMERIC","CALCULADO","ADX promedio del sector","Ver features_sector",""),
    ("features_ml","retorno_1d_sector_avg","NUMERIC","CALCULADO","Retorno 1d promedio del sector","Ver features_sector",""),
    ("features_ml","retorno_1d","NUMERIC","CALCULADO","Retorno futuro a 1 dia","close.shift(-1) / close - 1 * 100","Variable target auxiliar"),
    ("features_ml","retorno_5d","NUMERIC","CALCULADO","Retorno futuro a 5 dias","close.shift(-5) / close - 1 * 100",""),
    ("features_ml","retorno_10d","NUMERIC","CALCULADO","Retorno futuro a 10 dias","close.shift(-10) / close - 1 * 100",""),
    ("features_ml","retorno_20d","NUMERIC","CALCULADO","Retorno futuro a 20 dias (OBJETIVO PRINCIPAL)","close.shift(-20) / close - 1 * 100","Horizonte de prediccion: 1 mes habil"),
    ("features_ml","label","VARCHAR","CALCULADO","Clasificacion del retorno futuro 20d","GANANCIA si retorno_20d > 1% | PERDIDA si < -1% | NEUTRO si entre -1% y +1%","Umbral simetrico de +/- 1%"),
    ("features_ml","label_binario","SMALLINT","CALCULADO","Variable objetivo binaria","1 si label='GANANCIA' | 0 si PERDIDA o NEUTRO","Variable que predice el modelo ML"),

    # ── resultados_modelos_ml ─────────────────────────────────────────
    ("resultados_modelos_ml","scope","VARCHAR","CALCULADO","Alcance del modelo","'global' = todos los sectores | nombre_sector = modelo sectorial",""),
    ("resultados_modelos_ml","algoritmo","VARCHAR","CALCULADO","Algoritmo de ML","'rf' = Random Forest | 'xgb' = XGBoost | 'lgbm' = LightGBM",""),
    ("resultados_modelos_ml","segmento","VARCHAR","CALCULADO","Particion evaluada","TRAIN / TEST / BACKTEST","TEST es el segmento de comparacion principal"),
    ("resultados_modelos_ml","modelo_version","VARCHAR","CALCULADO","Version del feature set","'v1' = 29 feat | 'v2' = 59 feat (V1+PA) | 'v3' = 53 feat (V1+MS)",""),
    ("resultados_modelos_ml","n_filas","INTEGER","CALCULADO","Cantidad de muestras evaluadas","len(X_segmento)",""),
    ("resultados_modelos_ml","n_features","INTEGER","CALCULADO","Cantidad de features del modelo","29 (v1) / 59 (v2) / 53 (v3)",""),
    ("resultados_modelos_ml","accuracy","NUMERIC","CALCULADO","Exactitud global","Porcentaje de predicciones correctas sobre total","Puede ser enganoso con clases desbalanceadas"),
    ("resultados_modelos_ml","precision_w","NUMERIC","CALCULADO","Precision ponderada por clase","weighted avg precision de sklearn",""),
    ("resultados_modelos_ml","recall_w","NUMERIC","CALCULADO","Recall ponderado por clase","weighted avg recall de sklearn",""),
    ("resultados_modelos_ml","f1_w","NUMERIC","CALCULADO","F1 score ponderado","weighted avg F1 de sklearn",""),
    ("resultados_modelos_ml","precision_1","NUMERIC","CALCULADO","Precision de la clase GANANCIA (label=1)","TP / (TP + FP) para clase 1","De cada 10 predicciones GANANCIA, cuantas fueron correctas"),
    ("resultados_modelos_ml","recall_1","NUMERIC","CALCULADO","Recall de la clase GANANCIA (label=1)","TP / (TP + FN) para clase 1","De todas las GANANCIAS reales, cuantas el modelo detecto"),
    ("resultados_modelos_ml","f1_1","NUMERIC","CALCULADO","F1 de la clase GANANCIA (METRICA PRINCIPAL)","2 * precision_1 * recall_1 / (precision_1 + recall_1)","Equilibrio entre precision y recall para GANANCIA"),
    ("resultados_modelos_ml","roc_auc","NUMERIC","CALCULADO","Area bajo la curva ROC","sklearn.roc_auc_score(y_true, y_proba)","1.0 = perfecto | 0.5 = aleatorio"),

    # ── modelos_produccion ────────────────────────────────────────────
    ("modelos_produccion","scope","VARCHAR","CALCULADO","Alcance del modelo desplegado","'global' | nombre_sector | 'Automotive'",""),
    ("modelos_produccion","tipo","VARCHAR","CALCULADO","Tipo de modelo ganador","'global' = usa global champion | 'sectorial' = usa modelo propio del sector","Resultado del Challenger Final (Nivel 3)"),
    ("modelos_produccion","algoritmo","VARCHAR","CALCULADO","Algoritmo del modelo desplegado","rf / xgb / lgbm (o global_rf etc)","Ganador del challenger"),
    ("modelos_produccion","modelo_path","VARCHAR","CALCULADO","Ruta al archivo .joblib del modelo","models/ o models_v2/ o models_v3/ segun version","Serializado con joblib"),
    ("modelos_produccion","modelo_version","VARCHAR","CALCULADO","Version del modelo desplegado","v1 / v2 / v3",""),
    ("modelos_produccion","n_features","INTEGER","CALCULADO","Cantidad de features del modelo","29 / 59 / 53",""),
    ("modelos_produccion","f1_test","NUMERIC","CALCULADO","F1_GANANCIA en TEST del modelo desplegado","F1 clase 1 sobre segmento TEST","Metrica de seleccion del challenger"),
    ("modelos_produccion","f1_backtest","NUMERIC","CALCULADO","F1_GANANCIA en BACKTEST","F1 clase 1 sobre segmento BACKTEST","Validacion adicional"),
    ("modelos_produccion","roc_auc_test","NUMERIC","CALCULADO","ROC-AUC en TEST","Area bajo curva ROC en segmento TEST",""),
    ("modelos_produccion","activo","BOOLEAN","CALCULADO","Flag: modelo activo en produccion","TRUE por defecto al insertar",""),

    # ── features_precio_accion ────────────────────────────────────────
    ("features_precio_accion","ticker","VARCHAR","FUENTE","Simbolo bursatil","",""),
    ("features_precio_accion","fecha","DATE","FUENTE","Fecha de la barra","",""),
    ("features_precio_accion","body_pct","NUMERIC","CALCULADO","Retorno intradía firmado (%)","(close - open) / open * 100","Positivo = vela alcista"),
    ("features_precio_accion","body_ratio","NUMERIC","CALCULADO","Tamano del cuerpo respecto al rango total","|close - open| / (high - low)","Entre 0 y 1; 0 = doji puro, 1 = marubozu"),
    ("features_precio_accion","upper_shadow_pct","NUMERIC","CALCULADO","Sombra superior como fraccion del rango","(high - max(open,close)) / (high - low)","Entre 0 y 1"),
    ("features_precio_accion","lower_shadow_pct","NUMERIC","CALCULADO","Sombra inferior como fraccion del rango","(min(open,close) - low) / (high - low)","Entre 0 y 1"),
    ("features_precio_accion","es_alcista","SMALLINT","CALCULADO","Vela alcista (1) o bajista (0)","1 si close > open, 0 si no","Binario"),
    ("features_precio_accion","gap_apertura_pct","NUMERIC","CALCULADO","Gap de apertura respecto al cierre anterior (%)","(open - prev_close) / prev_close * 100","Positivo = gap alcista"),
    ("features_precio_accion","rango_diario_pct","NUMERIC","CALCULADO","Rango diario como % del cierre","(high - low) / close * 100","Mide amplitud de la sesion"),
    ("features_precio_accion","rango_rel_atr","NUMERIC","CALCULADO","Rango diario en multiples de ATR14","(high - low) / atr14","1 = rango normal; >2 = sesion amplia"),
    ("features_precio_accion","clv","NUMERIC","CALCULADO","Close Location Value","((close-low) - (high-close)) / (high-low)","Entre -1 y +1; +1 = cierre en maximo"),
    ("features_precio_accion","patron_doji","SMALLINT","CALCULADO","Patron Doji","1 si body_ratio < 0.05","Indecision: cuerpo muy pequeno"),
    ("features_precio_accion","patron_hammer","SMALLINT","CALCULADO","Patron Hammer (martillo)","1 si lower_shadow>60% AND body<30% AND es_alcista","Potencial reversion alcista desde soporte"),
    ("features_precio_accion","patron_shooting_star","SMALLINT","CALCULADO","Patron Shooting Star (estrella fugaz)","1 si upper_shadow>60% AND body<30% AND NOT es_alcista","Potencial reversion bajista desde resistencia"),
    ("features_precio_accion","patron_marubozu","SMALLINT","CALCULADO","Patron Marubozu (vela de fuerza)","1 si body_ratio > 0.85","Vela sin sombras, fuerza unidireccional"),
    ("features_precio_accion","patron_engulfing_bull","SMALLINT","CALCULADO","Patron Engulfing Alcista","1 si cuerpo actual > cuerpo previo AND vela alcista AND vela previa bajista","Envolvente alcista: patron de 2 velas"),
    ("features_precio_accion","patron_engulfing_bear","SMALLINT","CALCULADO","Patron Engulfing Bajista","1 si cuerpo actual > cuerpo previo AND vela bajista AND vela previa alcista","Envolvente bajista: patron de 2 velas"),
    ("features_precio_accion","inside_bar","SMALLINT","CALCULADO","Barra interior (inside bar)","1 si high < prev_high AND low > prev_low","Compresion de rango; potencial expansion futura"),
    ("features_precio_accion","outside_bar","SMALLINT","CALCULADO","Barra exterior (outside bar)","1 si high > prev_high AND low < prev_low","Rango que engloba la barra anterior"),
    ("features_precio_accion","body_pct_ma5","NUMERIC","CALCULADO","Media movil del body_pct en 5 dias","SMA(body_pct, 5)","Tendencia de corto plazo de velas"),
    ("features_precio_accion","velas_alcistas_5d","SMALLINT","CALCULADO","Conteo de velas alcistas en ultimas 5 barras","sum(es_alcista, 5)","Entre 0 y 5"),
    ("features_precio_accion","velas_alcistas_10d","SMALLINT","CALCULADO","Conteo de velas alcistas en ultimas 10 barras","sum(es_alcista, 10)","Entre 0 y 10"),
    ("features_precio_accion","rango_expansion","SMALLINT","CALCULADO","Expansion de rango inusual","1 si rango_diario_pct > 1.5 * SMA(rango_diario_pct, 10)","Volatilidad intradiaria elevada vs promedio"),
    ("features_precio_accion","dist_max_20d","NUMERIC","CALCULADO","Distancia al maximo de 20 dias (%)","(close - max_high_20d) / max_high_20d * 100","Negativo = precio bajo su maximo reciente"),
    ("features_precio_accion","dist_min_20d","NUMERIC","CALCULADO","Distancia al minimo de 20 dias (%)","(close - min_low_20d) / min_low_20d * 100","Positivo = precio sobre su minimo reciente"),
    ("features_precio_accion","pos_rango_20d","NUMERIC","CALCULADO","Posicion en el rango de 20 dias (0-1)","(close - min_close_20d) / (max_close_20d - min_close_20d)","0 = en el minimo del mes; 1 = en el maximo"),
    ("features_precio_accion","vol_ratio_5d","NUMERIC","CALCULADO","Volumen relativo vs promedio 5 dias","volume / SMA(volume, 5)",""),
    ("features_precio_accion","vol_spike","SMALLINT","CALCULADO","Pico de volumen inusual","1 si vol_relativo > 2.0","Volumen mas del doble del promedio"),
    ("features_precio_accion","up_vol_5d","NUMERIC","CALCULADO","Fraccion de volumen en dias alcistas (5d)","sum(volume * es_alcista, 5) / sum(volume, 5)","Entre 0 y 1; >0.6 = flujo comprador dominante"),
    ("features_precio_accion","chaikin_mf_20","NUMERIC","CALCULADO","Chaikin Money Flow de 20 dias","sum(CLV * volume, 20) / sum(volume, 20)","Entre -1 y +1; positivo = presion compradora"),
    ("features_precio_accion","vol_price_confirm","SMALLINT","CALCULADO","Confirmacion de volumen con movimiento de precio","1 si vol_spike=1 AND |body_pct| > media_body_abs_20d","Movimiento de precio acompanado de volumen alto"),
    ("features_precio_accion","vol_price_diverge","SMALLINT","CALCULADO","Divergencia volumen vs precio","1 si vol_spike=1 AND body_ratio < 0.10","Volumen alto pero vela pequena: posible distribucion"),

    # ── features_market_structure ─────────────────────────────────────
    ("features_market_structure","ticker","VARCHAR","FUENTE","Simbolo bursatil","",""),
    ("features_market_structure","fecha","DATE","FUENTE","Fecha de la barra","",""),
    ("features_market_structure","is_sh_5","SMALLINT","CALCULADO","Swing High con ventana N=5","1 si high[i] = max(high[i-5..i+5])","Rolling simetrico de 11 barras. Confirma pivot local"),
    ("features_market_structure","is_sl_5","SMALLINT","CALCULADO","Swing Low con ventana N=5","1 si low[i] = min(low[i-5..i+5])","Rolling simetrico de 11 barras"),
    ("features_market_structure","estructura_5","SMALLINT","CALCULADO","Estructura de mercado N=5","HH+HL -> +1 (alcista) | LH+LL -> -1 (bajista) | mixto -> 0","Requiere al menos 2 SH y 2 SL confirmados. HH=ultimo_SH > anterior_SH"),
    ("features_market_structure","dist_sh_5_pct","NUMERIC","CALCULADO","Distancia del cierre al ultimo Swing High N=5 (%)","(close - last_sh_5) / last_sh_5 * 100","Negativo = precio bajo el ultimo SH"),
    ("features_market_structure","dist_sl_5_pct","NUMERIC","CALCULADO","Distancia del cierre al ultimo Swing Low N=5 (%)","(close - last_sl_5) / last_sl_5 * 100","Positivo = precio sobre el ultimo SL"),
    ("features_market_structure","dias_sh_5","SMALLINT","CALCULADO","Barras transcurridas desde el ultimo Swing High N=5","posicion_actual - posicion_ultimo_SH","Max 252 barras (1 año habil)"),
    ("features_market_structure","dias_sl_5","SMALLINT","CALCULADO","Barras transcurridas desde el ultimo Swing Low N=5","posicion_actual - posicion_ultimo_SL","Max 252 barras"),
    ("features_market_structure","impulso_5_pct","NUMERIC","CALCULADO","Amplitud del swing actual N=5 (%)","|last_sh_5 - last_sl_5| / last_sl_5 * 100","Mide la magnitud del rango de swing actual"),
    ("features_market_structure","bos_bull_5","SMALLINT","CALCULADO","Break of Structure alcista N=5","1 si close cruza por encima del ultimo SH Y estructura_5 >= 0","BOS confirma la tendencia existente (alcista o neutral)"),
    ("features_market_structure","bos_bear_5","SMALLINT","CALCULADO","Break of Structure bajista N=5","1 si close cruza por debajo del ultimo SL Y estructura_5 <= 0","BOS confirma la tendencia existente (bajista o neutral)"),
    ("features_market_structure","choch_bull_5","SMALLINT","CALCULADO","Change of Character alcista N=5 (reversion alcista)","1 si close cruza por encima del ultimo SH Y estructura_5 < 0","CHoCH ocurre CONTRA la estructura: senial de reversion"),
    ("features_market_structure","choch_bear_5","SMALLINT","CALCULADO","Change of Character bajista N=5 (reversion bajista)","1 si close cruza por debajo del ultimo SL Y estructura_5 > 0","CHoCH ocurre CONTRA la estructura: senial de reversion"),
    ("features_market_structure","is_sh_10","SMALLINT","CALCULADO","Swing High con ventana N=10","1 si high[i] = max(high[i-10..i+10])","Rolling simetrico de 21 barras. Alineado con horizonte retorno_20d"),
    ("features_market_structure","is_sl_10","SMALLINT","CALCULADO","Swing Low con ventana N=10","1 si low[i] = min(low[i-10..i+10])","Rolling simetrico de 21 barras"),
    ("features_market_structure","estructura_10","SMALLINT","CALCULADO","Estructura de mercado N=10","HH+HL -> +1 | LH+LL -> -1 | mixto -> 0","Idem N=5 pero con ventana estrategica"),
    ("features_market_structure","dist_sh_10_pct","NUMERIC","CALCULADO","Distancia del cierre al ultimo Swing High N=10 (%)","(close - last_sh_10) / last_sh_10 * 100",""),
    ("features_market_structure","dist_sl_10_pct","NUMERIC","CALCULADO","Distancia del cierre al ultimo Swing Low N=10 (%)","(close - last_sl_10) / last_sl_10 * 100",""),
    ("features_market_structure","dias_sh_10","SMALLINT","CALCULADO","Barras desde el ultimo Swing High N=10","posicion_actual - posicion_ultimo_SH_10","Max 252. Feature mas importante en V3 (14.9% importance)"),
    ("features_market_structure","dias_sl_10","SMALLINT","CALCULADO","Barras desde el ultimo Swing Low N=10","posicion_actual - posicion_ultimo_SL_10","Max 252. 2da feature mas importante en V3 (12.6%)"),
    ("features_market_structure","impulso_10_pct","NUMERIC","CALCULADO","Amplitud del swing actual N=10 (%)","|last_sh_10 - last_sl_10| / last_sl_10 * 100","3ra feature mas importante en V3 (5.1%)"),
    ("features_market_structure","bos_bull_10","SMALLINT","CALCULADO","Break of Structure alcista N=10","1 si close cruza por encima del ultimo SH_10 Y estructura_10 >= 0","Igual que N=5 pero en escala estrategica"),
    ("features_market_structure","bos_bear_10","SMALLINT","CALCULADO","Break of Structure bajista N=10","1 si close cruza por debajo del ultimo SL_10 Y estructura_10 <= 0",""),
    ("features_market_structure","choch_bull_10","SMALLINT","CALCULADO","Change of Character alcista N=10","1 si close cruza por encima del ultimo SH_10 Y estructura_10 < 0","Reversion alcista de estructura estrategica"),
    ("features_market_structure","choch_bear_10","SMALLINT","CALCULADO","Change of Character bajista N=10","1 si close cruza por debajo del ultimo SL_10 Y estructura_10 > 0","Reversion bajista de estructura estrategica"),

    # ── resultados_ml_filter ──────────────────────────────────────────
    ("resultados_ml_filter","estrategia_entrada","VARCHAR","CALCULADO","Estrategia de entrada","E1-E4",""),
    ("resultados_ml_filter","estrategia_salida","VARCHAR","CALCULADO","Estrategia de salida","S1-S4",""),
    ("resultados_ml_filter","scope","VARCHAR","CALCULADO","Alcance del analisis","'GLOBAL' | nombre_sector",""),
    ("resultados_ml_filter","segmento","VARCHAR","CALCULADO","Particion temporal","TRAIN / TEST / BACKTEST",""),
    ("resultados_ml_filter","umbral_ml","NUMERIC","CALCULADO","Umbral de probabilidad del filtro ML","Modelo predice probabilidad de GANANCIA > umbral_ml para aprobar","Tipicamente 0.50"),
    ("resultados_ml_filter","ops_original","INTEGER","CALCULADO","Operaciones sin filtro ML (baseline)","COUNT de todas las operaciones de la estrategia en el segmento",""),
    ("resultados_ml_filter","win_rate_orig","NUMERIC","CALCULADO","Win rate sin filtro ML","ganancias / total sin filtro",""),
    ("resultados_ml_filter","ret_promedio_orig","NUMERIC","CALCULADO","Retorno promedio sin filtro","mean(retorno_pct) sin filtro",""),
    ("resultados_ml_filter","ret_total_orig","NUMERIC","CALCULADO","Retorno total sin filtro","sum(retorno_pct) sin filtro",""),
    ("resultados_ml_filter","profit_factor_orig","NUMERIC","CALCULADO","Profit factor sin filtro","sum(ganancias) / |sum(perdidas)| sin filtro",""),
    ("resultados_ml_filter","ops_ml","INTEGER","CALCULADO","Operaciones aprobadas por filtro ML","COUNT donde proba_ganancia > umbral_ml",""),
    ("resultados_ml_filter","ops_rechazadas","INTEGER","CALCULADO","Operaciones rechazadas por filtro ML","ops_original - ops_ml",""),
    ("resultados_ml_filter","pct_rechazo","NUMERIC","CALCULADO","Porcentaje de operaciones rechazadas","ops_rechazadas / ops_original",""),
    ("resultados_ml_filter","win_rate_ml","NUMERIC","CALCULADO","Win rate con filtro ML","ganancias_aprobadas / ops_ml",""),
    ("resultados_ml_filter","ret_promedio_ml","NUMERIC","CALCULADO","Retorno promedio con filtro ML","mean(retorno_pct) de operaciones aprobadas",""),
    ("resultados_ml_filter","ret_total_ml","NUMERIC","CALCULADO","Retorno total con filtro ML","sum(retorno_pct) de operaciones aprobadas",""),
    ("resultados_ml_filter","profit_factor_ml","NUMERIC","CALCULADO","Profit factor con filtro ML","sum(ganancias_aprobadas) / |sum(perdidas_aprobadas)|",""),
    ("resultados_ml_filter","delta_win_rate","NUMERIC","CALCULADO","Mejora en win rate (ML - Original)","win_rate_ml - win_rate_orig","Positivo = ML mejora la seleccion"),
    ("resultados_ml_filter","delta_ret_promedio","NUMERIC","CALCULADO","Mejora en retorno promedio","ret_promedio_ml - ret_promedio_orig",""),
    ("resultados_ml_filter","delta_profit_factor","NUMERIC","CALCULADO","Mejora en profit factor","profit_factor_ml - profit_factor_orig",""),
]


# ─────────────────────────────────────────────────────────────
# Estilos
# ─────────────────────────────────────────────────────────────

def _header_style():
    """Retorna estilo para celdas de encabezado."""
    return {
        "font":   Font(bold=True, color="FFFFFF", size=10),
        "fill":   PatternFill("solid", fgColor="1F3864"),
        "align":  Alignment(horizontal="center", vertical="center", wrap_text=True),
    }

def _subheader_style(color: str):
    return {
        "font":  Font(bold=True, color="FFFFFF", size=9),
        "fill":  PatternFill("solid", fgColor=color),
        "align": Alignment(horizontal="center", vertical="center"),
    }

def _apply_style(cell, style: dict):
    if "font"  in style: cell.font      = style["font"]
    if "fill"  in style: cell.fill      = style["fill"]
    if "align" in style: cell.alignment = style["align"]

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _formatear_hoja_datos(ws, df: pd.DataFrame, nombre_tabla: str):
    """Aplica formato basico a una hoja de datos."""
    # Encabezados
    hs = _header_style()
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _apply_style(cell, hs)
        cell.border = _thin_border()

    # Autoajuste de ancho (aprox)
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), 10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # Freeze primera fila
    ws.freeze_panes = "A2"


def _formatear_hoja_descripcion(ws):
    """Aplica formato a la hoja Descripcion."""
    colors = {
        "FUENTE":    "1A6B3C",
        "CALCULADO": "14456E",
    }

    col_widths = [28, 28, 12, 12, 50, 70, 55]
    col_letters = [get_column_letter(i) for i in range(1, 8)]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[col_letters[i]].width = w

    hs = _header_style()
    encabezados = ["Tabla", "Columna", "Tipo_BD", "Origen", "Descripcion", "Formula / Calculo", "Condiciones / Notas"]
    for col_idx, enc in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col_idx, value=enc)
        _apply_style(cell, hs)
        cell.border = _thin_border()

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    prev_tabla = None
    for row_idx, row_data in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
        tabla  = ws.cell(row=row_idx, column=1).value
        origen = ws.cell(row=row_idx, column=4).value
        color  = colors.get(origen or "", "374151")

        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if col_idx == 4 and origen in colors:
                cell.font = Font(bold=True, color="FFFFFF", size=9)
                cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif tabla != prev_tabla:
                if col_idx == 1:
                    cell.font = Font(bold=True, size=9, color="1F3864")
            else:
                if col_idx == 1:
                    cell.font = Font(color="888888", size=9)
                    cell.value = ""  # no repetir nombre de tabla

        prev_tabla = tabla
        ws.row_dimensions[row_idx].height = 42


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────

def main():
    inicio = datetime.now()
    ts     = inicio.strftime("%Y%m%d_%H%M")

    base_dir  = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    out_path  = os.path.join(base_dir, f"activos_ml_export_{ts}.xlsx")

    print("\n" + "=" * 65)
    print("  EXPORTAR BASE DE DATOS A EXCEL")
    print("=" * 65)
    print(f"  Tablas a exportar : {len(TABLAS)}")
    print(f"  Variables descritas: {len(DESC)}")
    print(f"  Destino           : {out_path}")
    print("=" * 65)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:

        # ── 1. Hoja Descripcion ───────────────────────────────────────
        print("\n  [1/N] Creando hoja Descripcion...")
        df_desc = pd.DataFrame(
            DESC,
            columns=["Tabla", "Columna", "Tipo_BD", "Origen",
                     "Descripcion", "Formula_Calculo", "Condiciones_Notas"]
        )
        df_desc.to_excel(writer, sheet_name="Descripcion", index=False)
        ws_desc = writer.sheets["Descripcion"]
        _formatear_hoja_descripcion(ws_desc)
        print(f"    {len(df_desc)} variables documentadas.")

        # ── 2. Hoja Resumen de tablas ──────────────────────────────────
        print("  [2/N] Creando hoja Resumen...")
        resumen_rows = []
        for tabla, descripcion in TABLAS:
            try:
                r = query_df(f"SELECT COUNT(*) AS n FROM {tabla}")
                n = int(r.iloc[0]["n"])
                resumen_rows.append({"Tabla": tabla, "Descripcion": descripcion, "Filas": n})
            except Exception:
                resumen_rows.append({"Tabla": tabla, "Descripcion": descripcion, "Filas": "ERROR"})

        df_res = pd.DataFrame(resumen_rows)
        df_res.to_excel(writer, sheet_name="Resumen", index=False)
        ws_res = writer.sheets["Resumen"]
        hs = _header_style()
        for col_idx in range(1, 4):
            cell = ws_res.cell(row=1, column=col_idx)
            _apply_style(cell, hs)
        ws_res.column_dimensions["A"].width = 30
        ws_res.column_dimensions["B"].width = 45
        ws_res.column_dimensions["C"].width = 12
        ws_res.freeze_panes = "A2"

        # ── 3. Una hoja por tabla ──────────────────────────────────────
        for i, (tabla, descripcion) in enumerate(TABLAS, 3):
            print(f"  [{i}/N] Exportando {tabla}...", end=" ", flush=True)
            try:
                df = query_df(f"SELECT * FROM {tabla} ORDER BY 1, 2 LIMIT 100000")
                nombre_hoja = tabla[:31]  # Excel max 31 chars
                df.to_excel(writer, sheet_name=nombre_hoja, index=False)
                ws = writer.sheets[nombre_hoja]
                _formatear_hoja_datos(ws, df, tabla)
                print(f"{len(df):,} filas x {len(df.columns)} cols")
            except Exception as e:
                print(f"ERROR: {e}")
                # Crear hoja vacia con mensaje de error
                df_err = pd.DataFrame({"Error": [str(e)]})
                df_err.to_excel(writer, sheet_name=tabla[:31], index=False)

    fin      = datetime.now()
    duracion = (fin - inicio).seconds
    print(f"\n{'='*65}")
    print(f"  Exportacion completada en {duracion}s")
    print(f"  Archivo: {out_path}")
    print("=" * 65)


if __name__ == "__main__":
    main()
